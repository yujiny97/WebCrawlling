import openpyxl
from pythoncode import webfunction_insta as w

#출처
#https://kimflstudio.tistory.com/entry/%ED%8C%8C%EC%9D%B4%EC%8D%AC-%EC%9D%B8%EC%8A%A4%ED%83%80%EA%B7%B8%EB%9E%A8-%ED%81%AC%EB%A1%A4%EB%A7%81-%EC%9D%B4%EB%AF%B8%EC%A7%80-%EB%8B%A4%EC%9A%B4%EB%A1%9C%EB%93%9C-beautifulsoup-selenium-%EC%82%AC%EC%9A%A9%EB%B2%95
##엑셀파일 경로
workbook = openpyxl.load_workbook('C:/Users/i5533/OneDrive/바탕 화면/바탕화면정리/빅데이터공모전/딥러닝/데이터엑셀/값.xlsx')
worksheet=workbook['Sheet1']#sheet이름

name_range=worksheet['D37':'D68'] #이름
i=37#start index
for row in name_range:
    for cell in row:
        name=cell.value
        print(name)
        newcnt=0
        curcnt=worksheet.cell(row=i,column=5).value
        w.getpicture(name, worksheet.cell(row=i, column=7).value,newcnt)
        #print(worksheet.cell(row=i,column=7).value) ##식용인지 독버섯인지
        i = i+1
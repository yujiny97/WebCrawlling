import openpyxl
from pythoncode import webfunction_naver as w

##엑셀파일 경로
workbook = openpyxl.load_workbook('C:/Users/i5533/OneDrive/바탕 화면/바탕화면정리/빅데이터공모전/딥러닝/데이터엑셀/값.xlsx')
worksheet=workbook['Sheet1']#sheet이름

name_range=worksheet['D40':'D66'] #이름
list_var=[]
i=40#start index
st='E'
for row in name_range:
    for cell in row:
        name=cell.value
        print(name)
        newcnt=0
        curcnt=worksheet.cell(row=i,column=5).value
        w.getpicture(name, worksheet.cell(row=i, column=7).value,newcnt)
        #print(worksheet.cell(row=i,column=7).value) ##식용인지 독버섯인지
        i = i+1